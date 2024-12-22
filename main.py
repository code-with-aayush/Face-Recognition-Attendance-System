import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime, date
from openpyxl import Workbook, load_workbook

path = 'Training_images'
images = []
classNames = []
myList = os.listdir(path)
print(myList)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
print(classNames)


def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList


def markAttendance(name):
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')

    file_name = "Attendance.xlsx"

    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Time", "Date"])

    entries = [(row[0].value, row[2].value) for row in sheet.iter_rows(min_row=2) if row[0].value and row[2].value]

    if (name, today_str) not in entries:
        now = datetime.now()
        time_str = now.strftime('%H:%M:%S')
        sheet.append([name, time_str, today_str])
        workbook.save(file_name)


encodeListKnown = findEncodings(images)
print('Encoding Complete')

cap = cv2.VideoCapture(1) #change it according to your system
# cap = cv2.VideoCapture(0) #this is your camera number

while True:
    success, img = cap.read()
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            name = classNames[matchIndex].upper()
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
            markAttendance(name)

    cv2.imshow('Webcam', img)
    cv2.waitKey(1)
